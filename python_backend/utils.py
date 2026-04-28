"""工具函数"""
import os
import re
import json
import tempfile
from copy import copy
from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.cell.cell import MergedCell

# Config
SHIPPER_NAME = 'AERO NEX FZCO'
SHIPPER_ADDR1 = 'No.529, 6W A, Dubai Airport Free Zone'
SHIPPER_ADDR2 = 'Dubai,United Arab Emirates'
COUNTRY_ORIGIN = 'Made in China'
ORIGIN = 'CHINA'

INV_DATA_START = 27
INV_RESERVED_ROWS = 7
PL_DATA_START = 27
PL_RESERVED_ROWS = 7

THIN = Side(style='thin')
NO_BORDER = Side(style=None)
BORDER_FULL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_NONE = Border(left=NO_BORDER, right=NO_BORDER, top=NO_BORDER, bottom=NO_BORDER)

TEMP_DIR = tempfile.mkdtemp(prefix='ci_pl_')
OUTPUT_DIR = os.path.join(TEMP_DIR, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format


def clear_row_safe(ws, row, start_col=1, end_col=8):
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None
        cell.border = BORDER_NONE


def set_full_border(ws, row, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        ws.cell(row=row, column=col_idx).border = BORDER_FULL


def clean_old_content(ws, start_row, end_row, start_col=1, end_col=8):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.border = BORDER_NONE


def map_cb_to_ean(cb_code, description, ean_map):
    if not cb_code or not cb_code.startswith('CB.'):
        return cb_code
    if not description or not ean_map:
        return cb_code
    desc_lower = description.lower().strip()
    for ean, desc in ean_map.items():
        if desc_lower == desc.lower().strip():
            return ean
    for ean, desc in ean_map.items():
        desc_ean_lower = desc.lower().strip()
        if desc_lower in desc_ean_lower or desc_ean_lower in desc_lower:
            return ean
    core = re.sub(r'\s*(with|w/)\s+.*$', '', desc_lower, flags=re.IGNORECASE)
    core = re.sub(r'\s*\(.*?\)$', '', core).strip()
    if core and len(core) > 3:
        for ean, desc in ean_map.items():
            desc_ean_lower = desc.lower().strip()
            desc_core = re.sub(r'\s*\(.*?\)$', '', desc_ean_lower).strip()
            if core in desc_core or desc_core in core:
                return ean
    return cb_code


def format_weight(weight_val):
    if not weight_val:
        return ''
    w_str = str(weight_val).strip()
    if 'KG' in w_str.upper():
        return w_str
    m = re.search(r'([0-9]+\.?[0-9]*)', w_str)
    if m:
        return f"{m.group(1)} KG"
    return w_str


def summarize_packages(ar_items):
    seen_pkgs = set()
    for item in ar_items:
        pkgs = str(item.get('pkgs', '')).strip().upper()
        if pkgs:
            seen_pkgs.add(pkgs)
    types = Counter()
    for pkgs in seen_pkgs:
        m = re.match(r'(BOX|PLT|CARTON|PALLET)[#\s]*(\d+)', pkgs, re.IGNORECASE)
        if m:
            pkg_type = m.group(1).upper()
        else:
            m = re.match(r'(\d+)\s*(BOX|PLT|CARTON|PALLET)', pkgs, re.IGNORECASE)
            if m:
                pkg_type = m.group(2).upper()
            else:
                continue
        if pkg_type in ['BOX', 'CARTON']:
            types['BOX'] += 1
        elif pkg_type in ['PLT', 'PALLET']:
            types['PLT'] += 1
    parts = []
    if types['BOX'] > 0:
        parts.append(f"{types['BOX']} BOX")
    if types['PLT'] > 0:
        parts.append(f"{types['PLT']} PLT")
    return ', '.join(parts)


def load_ean_map():
    ean_path = os.path.join(DATA_DIR, 'EAN 和货品对应关系.xlsx')
    df = pd.read_excel(ean_path)
    mapping = {}
    for _, row in df.iterrows():
        ean = str(row['EAN code']).strip()
        desc = str(row['Description']).strip()
        mapping[ean] = desc
    return mapping


def load_hs_code_map():
    hs_path = os.path.join(DATA_DIR, 'hs_code_map.json')
    with open(hs_path, 'r', encoding='utf-8') as f:
        return json.load(f)
