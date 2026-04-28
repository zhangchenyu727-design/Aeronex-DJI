"""Excel生成逻辑"""
import os
import re
import tempfile
from copy import copy
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.cell.cell import MergedCell

from utils import (
    SHIPPER_NAME, SHIPPER_ADDR1, SHIPPER_ADDR2, COUNTRY_ORIGIN, ORIGIN,
    INV_DATA_START, INV_RESERVED_ROWS, PL_DATA_START, PL_RESERVED_ROWS,
    THIN, NO_BORDER, BORDER_FULL, BORDER_NONE,
    copy_cell_style, clear_row_safe, set_full_border, clean_old_content,
    format_weight, DATA_DIR, OUTPUT_DIR
)

TEMPLATE_PATH = os.path.join(DATA_DIR, 'CI_PL_Template.xlsx')


def _unmerge_cell(ws, row, col):
    """如果指定单元格在合并范围内，解除该合并"""
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            ws.unmerge_cells(str(merged))
            break

def build_invoice(pi_data, ar_data, invoice_products, package_summary):
    template_wb = load_workbook(TEMPLATE_PATH)
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Invoice']
    template_ws = template_wb['Invoice']
    num_products = len(invoice_products)

    # 预先解除可能冲突的合并单元格
    for r in range(3, 11):
        for c in range(1, 7):
            _unmerge_cell(ws, r, c)

    if num_products <= INV_RESERVED_ROWS:
        end_row = INV_DATA_START + num_products - 1
        clear_row = INV_DATA_START + num_products
        for row in range(clear_row, INV_DATA_START + INV_RESERVED_ROWS):
            clean_old_content(ws, row, row)
        copy_cell_style(template_ws.cell(row=INV_DATA_START + INV_RESERVED_ROWS - 1, column=1),
                       ws.cell(row=clear_row, column=1))
        copy_cell_style(template_ws.cell(row=INV_DATA_START + INV_RESERVED_ROWS - 1, column=1),
                       ws.cell(row=end_row, column=1))

        for i, prod in enumerate(invoice_products):
            row = INV_DATA_START + i
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=1, value=prod.get('ean', ''))
            ws.cell(row=row, column=2, value=prod.get('description', ''))
            ws.cell(row=row, column=3, value=ORIGIN)
            ws.cell(row=row, column=4, value=prod.get('qty', 1))
            ws.cell(row=row, column=5, value=prod.get('rate', 0))
            ws.cell(row=row, column=6, value=prod.get('amount', 0))
            for col_idx in range(1, 9):
                ws.cell(row=row, column=col_idx).border = BORDER_FULL

        for col_idx in range(1, 9):
            ws.cell(row=end_row, column=col_idx).border = BORDER_FULL
    else:
        for i in range(INV_RESERVED_ROWS, num_products):
            ws.insert_rows(INV_DATA_START + i)
            src_row = INV_DATA_START + INV_RESERVED_ROWS - 1
            dst_row = INV_DATA_START + i
            for col_idx in range(1, 9):
                src_cell = template_ws.cell(row=src_row, column=col_idx)
                dst_cell = ws.cell(row=dst_row, column=col_idx)
                copy_cell_style(src_cell, dst_cell)

        for i, prod in enumerate(invoice_products):
            row = INV_DATA_START + i
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = Font(name='Times New Roman', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=1, value=prod.get('ean', ''))
            ws.cell(row=row, column=2, value=prod.get('description', ''))
            ws.cell(row=row, column=3, value=ORIGIN)
            ws.cell(row=row, column=4, value=prod.get('qty', 1))
            ws.cell(row=row, column=5, value=prod.get('rate', 0))
            ws.cell(row=row, column=6, value=prod.get('amount', 0))
            for col_idx in range(1, 9):
                ws.cell(row=row, column=col_idx).border = BORDER_FULL

    summary_row = INV_DATA_START + num_products
    for c in range(1, 9):
        _unmerge_cell(ws, summary_row, c)
    ws.cell(row=summary_row, column=1, value=f"=SUM(D{INV_DATA_START}:D{INV_DATA_START + num_products - 1})")
    ws.cell(row=summary_row, column=3, value="TOTAL AMOUNT")
    ws.cell(row=summary_row, column=6, value=f"=SUM(F{INV_DATA_START}:F{INV_DATA_START + num_products - 1})")
    ws.cell(row=summary_row, column=1).font = Font(name='Times New Roman', size=12, bold=True)
    ws.cell(row=summary_row, column=3).font = Font(name='Times New Roman', size=12, bold=True)
    ws.cell(row=summary_row, column=6).font = Font(name='Times New Roman', size=12, bold=True)
    set_full_border(ws, summary_row, 1, 8)

    ws.cell(row=3, column=2, value=pi_data.get('pi_number', ''))
    ws.cell(row=4, column=2, value=ar_data.get('date', ''))
    ws.cell(row=5, column=2, value=pi_data.get('pi_number', ''))
    ws.cell(row=5, column=6, value=pi_data.get('final_destination', ''))
    ws.cell(row=6, column=2, value=ar_data.get('order_number', ''))
    ws.cell(row=8, column=2, value=SHIPPER_NAME)
    ws.cell(row=9, column=2, value=SHIPPER_ADDR1)
    ws.cell(row=10, column=2, value=SHIPPER_ADDR2)

    bt_lines = [pi_data.get('bill_to_name', '')]
    if pi_data.get('bill_to_address'):
        bt_lines.append(pi_data.get('bill_to_address', ''))
    if pi_data.get('contact_name'):
        bt_lines.append(f"Contact: {pi_data['contact_name']}")
    if pi_data.get('email'):
        bt_lines.append(pi_data['email'])
    ws.cell(row=8, column=5, value='\n'.join(bt_lines))
    ws.cell(row=8, column=5).alignment = Alignment(wrap_text=True, vertical='top')

    ws.cell(row=10, column=6, value=COUNTRY_ORIGIN)

    if package_summary:
        for r in range(21, 25):
            val = ws.cell(row=r, column=1).value
            if isinstance(val, str) and 'TOTAL' in val and 'PACKAGES' in val:
                ws.cell(row=r, column=6, value=package_summary)
                break

    sig_row1 = summary_row + 3
    ws.cell(row=sig_row1, column=1, value="Saudi Customs")
    ws.cell(row=sig_row1 + 1, column=1, value="____________________________________")
    ws.cell(row=sig_row1 + 1, column=5, value="____________________________________")
    ws.cell(row=sig_row1 + 2, column=1, value="Customer")
    ws.cell(row=sig_row1 + 2, column=5, value="Exporter")
    ws.cell(row=sig_row1, column=1).font = Font(name='Times New Roman', size=11, italic=True)
    ws.cell(row=sig_row1 + 2, column=1).font = Font(name='Times New Roman', size=11, italic=True)
    ws.cell(row=sig_row1 + 2, column=5).font = Font(name='Times New Roman', size=11, italic=True)

    return wb


def _write_pl_row(ws, row, prod):
    """写入PL数据行，统一设置整行格式（字体/对齐/边框）。
    WEIGHT/DIMENSION 由调用方根据 PKGS 分组进行合并处理。"""
    for col_idx in range(1, 9):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(name='Times New Roman', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_FULL
    ws.cell(row=row, column=1, value=prod.get('hs_code', ''))
    ws.cell(row=row, column=2, value=prod.get('ean', ''))
    ws.cell(row=row, column=3, value=prod.get('pi_number', ''))
    ws.cell(row=row, column=4, value=prod.get('description', ''))
    ws.cell(row=row, column=5, value=ORIGIN)
    ws.cell(row=row, column=6, value=prod.get('qty', 1))


def build_packing_list(pi_data, ar_data, pl_items):
    template_wb = load_workbook(TEMPLATE_PATH)
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Packing List']
    template_ws = template_wb['Packing List']
    num_items = len(pl_items)

    # 预先解除可能冲突的合并单元格
    for r in range(3, 14):
        for c in range(1, 7):
            _unmerge_cell(ws, r, c)

    ws.cell(row=3, column=2, value=pi_data.get('pi_number', ''))
    ws.cell(row=4, column=2, value=ar_data.get('date', ''))
    ws.cell(row=5, column=2, value=pi_data.get('pi_number', ''))
    ws.cell(row=5, column=6, value=pi_data.get('final_destination', ''))
    ws.cell(row=6, column=2, value=ar_data.get('order_number', ''))
    ws.cell(row=8, column=2, value=SHIPPER_NAME)
    ws.cell(row=9, column=2, value=SHIPPER_ADDR1)
    ws.cell(row=10, column=2, value=SHIPPER_ADDR2)

    bt_lines = [pi_data.get('bill_to_name', '')]
    if pi_data.get('bill_to_address'):
        bt_lines.append(pi_data.get('bill_to_address', ''))
    if pi_data.get('contact_name'):
        bt_lines.append(f"Contact: {pi_data['contact_name']}")
    if pi_data.get('email'):
        bt_lines.append(pi_data['email'])
    ws.cell(row=8, column=5, value='\n'.join(bt_lines))
    ws.cell(row=8, column=5).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=10, column=6, value=COUNTRY_ORIGIN)

    if num_items <= PL_RESERVED_ROWS:
        end_row = PL_DATA_START + num_items - 1
        clear_row = PL_DATA_START + num_items
        for row in range(clear_row, PL_DATA_START + PL_RESERVED_ROWS):
            clean_old_content(ws, row, row)
        copy_cell_style(template_ws.cell(row=PL_DATA_START + PL_RESERVED_ROWS - 1, column=1),
                       ws.cell(row=clear_row, column=1))
        copy_cell_style(template_ws.cell(row=PL_DATA_START + PL_RESERVED_ROWS - 1, column=1),
                       ws.cell(row=end_row, column=1))

        for i, prod in enumerate(pl_items):
            row = PL_DATA_START + i
            _write_pl_row(ws, row, prod)
            is_first = (i == 0) or (pl_items[i-1].get('pkgs') != prod.get('pkgs'))
            if is_first:
                ws.cell(row=row, column=7, value=format_weight(prod.get('weight', '')))
                ws.cell(row=row, column=8, value=prod.get('dimension', ''))
    else:
        for i in range(PL_RESERVED_ROWS, num_items):
            ws.insert_rows(PL_DATA_START + i)
            src_row = PL_DATA_START + PL_RESERVED_ROWS - 1
            dst_row = PL_DATA_START + i
            for col_idx in range(1, 9):
                src_cell = template_ws.cell(row=src_row, column=col_idx)
                dst_cell = ws.cell(row=dst_row, column=col_idx)
                copy_cell_style(src_cell, dst_cell)

        for i, prod in enumerate(pl_items):
            row = PL_DATA_START + i
            _write_pl_row(ws, row, prod)
            is_first = (i == 0) or (pl_items[i-1].get('pkgs') != prod.get('pkgs'))
            if is_first:
                ws.cell(row=row, column=7, value=format_weight(prod.get('weight', '')))
                ws.cell(row=row, column=8, value=prod.get('dimension', ''))

    # 按 PKGS 分组重建 WEIGHT(G) 和 DIMENSION(H) 列的合并单元格
    current_pkg = None
    start_idx = None
    for i, prod in enumerate(pl_items):
        pkg = prod.get('pkgs', '')
        if pkg != current_pkg:
            if current_pkg is not None and start_idx is not None and i - 1 > start_idx:
                start_row = PL_DATA_START + start_idx
                end_row = PL_DATA_START + i - 1
                ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)
                ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)
            current_pkg = pkg
            start_idx = i

    if current_pkg is not None and start_idx is not None and len(pl_items) - 1 > start_idx:
        start_row = PL_DATA_START + start_idx
        end_row = PL_DATA_START + len(pl_items) - 1
        ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)
        ws.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)

    summary_row = PL_DATA_START + num_items
    for c in range(1, 9):
        _unmerge_cell(ws, summary_row, c)
    ws.cell(row=summary_row, column=1, value=f"=SUM(F{PL_DATA_START}:F{PL_DATA_START + num_items - 1})")
    ws.cell(row=summary_row, column=3, value="TOTAL WEIGHT")
    ws.cell(row=summary_row, column=6, value=f"=SUM(F{PL_DATA_START}:F{PL_DATA_START + num_items - 1})")
    ws.cell(row=summary_row, column=1).font = Font(name='Times New Roman', size=12, bold=True)
    ws.cell(row=summary_row, column=3).font = Font(name='Times New Roman', size=12, bold=True)
    ws.cell(row=summary_row, column=6).font = Font(name='Times New Roman', size=12, bold=True)
    set_full_border(ws, summary_row, 1, 8)

    if ar_data.get('consignee'):
        ws.cell(row=13, column=1, value=f"Consignee: {ar_data['consignee']}")

    sig_row1 = summary_row + 3
    ws.cell(row=sig_row1, column=1, value="Saudi Customs")
    ws.cell(row=sig_row1 + 1, column=1, value="____________________________________")
    ws.cell(row=sig_row1 + 1, column=5, value="____________________________________")
    ws.cell(row=sig_row1 + 2, column=1, value="Customer")
    ws.cell(row=sig_row1 + 2, column=5, value="Exporter")
    ws.cell(row=sig_row1, column=1).font = Font(name='Times New Roman', size=11, italic=True)
    ws.cell(row=sig_row1 + 2, column=1).font = Font(name='Times New Roman', size=11, italic=True)
    ws.cell(row=sig_row1 + 2, column=5).font = Font(name='Times New Roman', size=11, italic=True)

    return wb
